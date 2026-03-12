import csv
import io
from typing import List, Dict, Tuple
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from app.models.user import User
from app.models.skill import Skill, UserSkill
from app.models.enums import UserRole
from app.utils.security import get_password_hash


MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
MAX_ROWS = 5000
REQUIRED_COLUMNS = {"email", "full_name"}
OPTIONAL_COLUMNS = {"department", "password", "skills"}


def parse_xlsx(file_content: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse an xlsx file and return rows + errors."""
    errors = []
    rows = []

    try:
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]

        # Validate required columns
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            return [], [f"Missing required columns: {', '.join(missing)}"]

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if i > MAX_ROWS + 1:
                errors.append(f"File exceeds maximum of {MAX_ROWS} rows. Only first {MAX_ROWS} rows processed.")
                break

            row_dict = {}
            for j, header in enumerate(headers):
                if header and j < len(row):
                    row_dict[header] = str(row[j]).strip() if row[j] is not None else ""

            if row_dict.get("email") and row_dict.get("full_name"):
                rows.append(row_dict)
            elif any(row_dict.values()):
                errors.append(f"Row {i}: Missing required fields (email or full_name)")

        wb.close()
    except Exception as e:
        return [], [f"Error parsing xlsx file: {str(e)}"]

    return rows, errors


def parse_csv(file_content: bytes) -> Tuple[List[Dict], List[str]]:
    """Parse a CSV file and return rows + errors."""
    errors = []
    rows = []

    try:
        text = file_content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))

        headers = set(h.strip().lower() for h in (reader.fieldnames or []))
        missing = REQUIRED_COLUMNS - headers
        if missing:
            return [], [f"Missing required columns: {', '.join(missing)}"]

        for i, row in enumerate(reader, start=2):
            if i > MAX_ROWS + 1:
                errors.append(f"File exceeds maximum of {MAX_ROWS} rows. Only first {MAX_ROWS} rows processed.")
                break

            # Normalize keys
            row_dict = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}

            if row_dict.get("email") and row_dict.get("full_name"):
                rows.append(row_dict)
            elif any(row_dict.values()):
                errors.append(f"Row {i}: Missing required fields (email or full_name)")

    except Exception as e:
        return [], [f"Error parsing CSV file: {str(e)}"]

    return rows, errors


def import_employees(
    db: Session, rows: List[Dict], default_password: str = "changeme123"
) -> Tuple[int, int, List[str]]:
    """Import employee records from parsed rows.
    Returns (created_count, skipped_count, errors)."""
    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=1):
        email = row.get("email", "").lower().strip()
        full_name = row.get("full_name", "").strip()

        if not email or not full_name:
            errors.append(f"Row {i}: Missing email or full_name")
            skipped += 1
            continue

        # Check if user already exists
        existing = db.query(User).filter(User.email == email).first()
        if existing:
            errors.append(f"Row {i}: Email '{email}' already exists — skipped")
            skipped += 1
            continue

        try:
            password = row.get("password", "").strip() or default_password
            user = User(
                email=email,
                full_name=full_name,
                hashed_password=get_password_hash(password),
                role=UserRole.EMPLOYEE,
                department=row.get("department", "").strip() or None,
            )
            db.add(user)
            db.flush()

            # Handle skills (comma-separated skill names)
            skills_str = row.get("skills", "").strip()
            if skills_str:
                skill_names = [s.strip() for s in skills_str.split(",") if s.strip()]
                for skill_name in skill_names:
                    # Find or create skill
                    skill = db.query(Skill).filter(Skill.name == skill_name).first()
                    if not skill:
                        skill = Skill(name=skill_name)
                        db.add(skill)
                        db.flush()

                    user_skill = UserSkill(user_id=user.id, skill_id=skill.id, proficiency_level=3)
                    db.add(user_skill)

            created += 1
        except Exception as e:
            errors.append(f"Row {i}: Error importing '{email}' — {str(e)}")
            skipped += 1
            db.rollback()
            continue

    db.commit()
    return created, skipped, errors


def process_upload(db: Session, filename: str, file_content: bytes) -> Dict:
    """Process an uploaded employee spreadsheet (xlsx or csv)."""
    if len(file_content) > MAX_FILE_SIZE:
        return {"success": False, "message": "File exceeds 5MB limit", "created": 0, "skipped": 0, "errors": []}

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "xlsx":
        rows, parse_errors = parse_xlsx(file_content)
    elif ext == "csv":
        rows, parse_errors = parse_csv(file_content)
    else:
        return {"success": False, "message": "Unsupported file format. Use .xlsx or .csv", "created": 0, "skipped": 0, "errors": []}

    if parse_errors and not rows:
        return {"success": False, "message": "Failed to parse file", "created": 0, "skipped": 0, "errors": parse_errors}

    created, skipped, import_errors = import_employees(db, rows)
    all_errors = parse_errors + import_errors

    return {
        "success": True,
        "message": f"Import complete: {created} created, {skipped} skipped",
        "created": created,
        "skipped": skipped,
        "errors": all_errors,
    }
